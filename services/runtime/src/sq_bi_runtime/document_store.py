"""Document ingestion: upload, extract field/term hints, feed evidence.

Supported formats: Excel (.xlsx/.xls), CSV (.csv), text (.txt/.md),
Word (.docx), PDF (.pdf).  All parsing failures are logged and degrade
gracefully — the profile builds from metadata alone if no documents are
processable.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class FieldHint:
    """A term/field hint extracted from a document."""
    term: str
    synonyms: list[str] = field(default_factory=list)
    description: str | None = None
    source_document: str | None = None


@dataclass
class DocumentExtractionResult:
    """Outcome of parsing one uploaded file."""
    document_id: str
    filename: str
    hints: list[FieldHint]
    success: bool
    error: str | None = None


class DocumentStore:
    """Extracts field/term hints from uploaded data-dictionary files.

    The store writes the extracted hints to the semantic profile store as
    evidence entries; it never blocks a scan.
    """

    def __init__(self, storage_dir: str | Path) -> None:
        self._storage_dir = Path(storage_dir)
        try:
            self._storage_dir.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            raise RuntimeError(
                f"Cannot create document storage directory {self._storage_dir}: {exc}"
            ) from exc

    def save_file(self, document_id: str, filename: str, content: bytes) -> Path:
        """Persist raw bytes; returns the saved path."""
        dest = self._storage_dir / f"{document_id}_{filename}"
        dest.write_bytes(content)
        return dest

    def delete_file(self, document_id: str, filename: str) -> None:
        """Remove the persisted bytes for a document, if present."""
        dest = self._storage_dir / f"{document_id}_{filename}"
        dest.unlink(missing_ok=True)

    def extract_hints(
        self, document_id: str, filename: str, content: bytes
    ) -> DocumentExtractionResult:
        """Extract field/term hints from file content.

        Returns a result even on failure (success=False, hints=[]).
        """
        suffix = Path(filename).suffix.lower()
        try:
            if suffix in (".xlsx", ".xls"):
                hints = self._extract_from_excel(document_id, filename, content)
            elif suffix == ".csv":
                hints = self._extract_from_csv(document_id, filename, content)
            elif suffix in (".txt", ".md"):
                hints = self._extract_from_text(document_id, filename, content)
            elif suffix == ".docx":
                hints = self._extract_from_docx(document_id, filename, content)
            elif suffix == ".pdf":
                hints = self._extract_from_pdf(document_id, filename, content)
            else:
                return DocumentExtractionResult(
                    document_id=document_id,
                    filename=filename,
                    hints=[],
                    success=False,
                    error=f"Unsupported format: {suffix}",
                )
            return DocumentExtractionResult(
                document_id=document_id,
                filename=filename,
                hints=hints,
                success=True,
            )
        except Exception as exc:
            logger.warning(
                "document_store.extract.failed",
                extra={"document_id": document_id, "filename": filename, "error": str(exc)},
            )
            return DocumentExtractionResult(
                document_id=document_id,
                filename=filename,
                hints=[],
                success=False,
                error=str(exc),
            )

    # ── Extractors ────────────────────────────────────────────────────

    def _extract_from_excel(
        self, document_id: str, filename: str, content: bytes
    ) -> list[FieldHint]:
        try:
            import openpyxl  # type: ignore[import]
        except ImportError:
            raise RuntimeError("openpyxl not installed; cannot parse Excel files.")
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        hints: list[FieldHint] = []
        for ws in wb.worksheets:
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(), [])]
            term_col = _find_column(headers, ["字段名", "field_name", "column_name", "字段", "列名"])
            desc_col = _find_column(headers, ["说明", "description", "备注", "注释", "comment"])
            syn_col = _find_column(headers, ["别名", "synonym", "synonyms", "同义词"])
            for row in ws.iter_rows(min_row=2, values_only=True):
                if term_col is None or term_col >= len(row):
                    continue
                term = str(row[term_col] or "").strip()
                if not term:
                    continue
                desc = str(row[desc_col] or "").strip() if desc_col is not None and desc_col < len(row) else None
                syn_raw = str(row[syn_col] or "").strip() if syn_col is not None and syn_col < len(row) else ""
                synonyms = [s.strip() for s in syn_raw.split(",") if s.strip()] if syn_raw else []
                hints.append(FieldHint(
                    term=term.upper(),
                    synonyms=synonyms,
                    description=desc or None,
                    source_document=filename,
                ))
        return hints

    def _extract_from_csv(
        self, document_id: str, filename: str, content: bytes
    ) -> list[FieldHint]:
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        fieldnames = reader.fieldnames or []
        term_key = _find_dict_key(fieldnames, ["字段名", "field_name", "column_name", "字段", "列名"])
        desc_key = _find_dict_key(fieldnames, ["说明", "description", "备注", "注释", "comment"])
        syn_key = _find_dict_key(fieldnames, ["别名", "synonym", "synonyms", "同义词"])
        hints: list[FieldHint] = []
        for row in reader:
            if not term_key:
                break
            term = str(row.get(term_key) or "").strip().upper()
            if not term:
                continue
            desc = str(row.get(desc_key) or "").strip() if desc_key else None
            syn_raw = str(row.get(syn_key) or "").strip() if syn_key else ""
            synonyms = [s.strip() for s in syn_raw.split(",") if s.strip()] if syn_raw else []
            hints.append(FieldHint(
                term=term,
                synonyms=synonyms,
                description=desc or None,
                source_document=filename,
            ))
        return hints

    def _extract_from_text(
        self, document_id: str, filename: str, content: bytes
    ) -> list[FieldHint]:
        """Extract any UPPER_CASE or snake_case identifiers as potential field hints."""
        import re
        text = content.decode("utf-8", errors="replace")
        identifiers = re.findall(r'\b[A-Z][A-Z0-9_]{2,}\b', text)
        seen: set[str] = set()
        hints: list[FieldHint] = []
        for ident in identifiers:
            if ident not in seen:
                seen.add(ident)
                hints.append(FieldHint(term=ident, source_document=filename))
        return hints

    def _extract_from_docx(
        self, document_id: str, filename: str, content: bytes
    ) -> list[FieldHint]:
        try:
            from docx import Document  # type: ignore[import]
        except ImportError:
            raise RuntimeError("python-docx not installed; cannot parse Word files.")
        doc = Document(io.BytesIO(content))
        lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(lines)
        return self._extract_from_text(document_id, filename, text.encode())

    def _extract_from_pdf(
        self, document_id: str, filename: str, content: bytes
    ) -> list[FieldHint]:
        try:
            import pdfplumber  # type: ignore[import]
        except ImportError:
            raise RuntimeError("pdfplumber not installed; cannot parse PDF files.")
        text_parts: list[str] = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return self._extract_from_text(document_id, filename, "\n".join(text_parts).encode())


# ── Helpers ───────────────────────────────────────────────────────────

def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    lower_headers = [h.lower() for h in headers]
    for candidate in candidates:
        try:
            return lower_headers.index(candidate.lower())
        except ValueError:
            pass
    return None


def _find_dict_key(keys: list[str], candidates: list[str]) -> str | None:
    lower_keys = {k.lower(): k for k in keys}
    for candidate in candidates:
        if candidate.lower() in lower_keys:
            return lower_keys[candidate.lower()]
    return None
